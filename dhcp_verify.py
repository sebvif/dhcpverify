import paramiko
import time
import urllib
import os.path
import openpyxl
import ftplib
import sys
import ipaddress
import netaddr

class bcolors:
	HEADER = '\033[95m'
	OKBLUE = '\033[94m'
	OKGREEN = '\033[92m'
	WARNING = '\033[93m'
	FAIL = '\033[91m'
	ENDC = '\033[0m'
	BOLD = '\033[1m'
	UNDERLINE = '\033[4m'

def nextMAC(mac):
	mac=netaddr.EUI(mac)
	next_mac=netaddr.EUI(int(mac)+1)
	next_mac.dialect=netaddr.mac_unix_expanded
	return str(next_mac).upper()

def mapColNum(col):
	col=col.lower()
	letters=list(col)
	if len(letters)>1:
		col_num=((int(ord(letters[0]))-96)*26)+(int(ord(letters[1]))-96)
		return col_num
	else:
		col_num=int(ord(letters[0]))-96
		return col_num

def getData(db,sheet):
	wb=openpyxl.load_workbook(db)
	data=wb[sheet]
	return data

def formatMACaddress(mac):
	mac=str(mac)
	format_mac=""
	if ':' in mac:
		return mac
	else:
		for i in range(len(mac)):
			if i%2==0 or i==len(mac)-1:
				format_mac=format_mac+mac[i]
			elif i%2==1:
				format_mac=format_mac+mac[i]+':'
		return format_mac

def getNetworks(db,sheet):
	wb=openpyxl.load_workbook(db)
	data=wb[sheet]
	i=2
	networks={}
	while i<data.max_row:
		#print i
		#print data.max_row
		current_site=data.cell(i,mapColNum('A')).value
		next_site=data.cell(i+1,mapColNum('A')).value
		if current_site!=next_site:
			networks[current_site]=data.cell(i,mapColNum('B')).value
		i+=1
	return networks

def remote_copy_to_tmp_dir(svr,port,file_path,dest_path):
	ssh=paramiko.SSHClient()
	ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
	try:
		ssh.connect(svr,port=port,username="cwmp",password="!El4r4")
		command='cp '+file_path+' '+dest_path+time.strftime("%Y%m%d")
		stdin, stdout, stderr = ssh.exec_command(command)
		return stdin,stdout,stderr
	except Exception, e:
		print bcolors.FAIL+'Error en la conexion al servidor.'+bcolors.ENDC
		return False

def ftp_download(svr,src_file,dst_path):
	path='ftp://cwmp:!El4r4@'+svr+src_file
	try:
		#print path
		urllib.urlretrieve(path,dst_path)
		return os.path.isfile(dst_path)#True
	except Exception, e:
		print bcolors.FAIL+'Error en la descarga del archivo remoto.'+bcolors.ENDC
		#print e
		return False

def ftp_upload(svr,src_file,dst_file):
	session = ftplib.FTP('189.240.5.86','cwmp','!El4r4')
	file = open(src_file,'rb')                  # file to send
	try:
		print dst_file
		session.storbinary('STOR '+dst_file, file)     # send the file
		file.close()                                    # close file and FTP
		session.quit()
		print bcolors.OKGREEN+'Archivo subido con exito.'+bcolors.ENDC
		return True
	except Exception, e:
		print e
		print bcolors.FAIL+'Error en la carga del archivo.'+bcolors.ENDC
		return False

def check_macs_in_file(file_name,macs):
#en esta funcion falta revisar el # del comentario
	old_macs=[]
	new_macs=[]
	result={}
	found_flag=False
	for mac in macs:
		for line in file(file_name):
			if mac in line:
				print bcolors.WARNING+"Direccion MAC encontrada: "+bcolors.ENDC+bcolors.UNDERLINE+mac+bcolors.ENDC
				found_flag=True
				break
		if found_flag:
			old_macs.append(mac)
		else:
			new_macs.append(mac)
	result.update({'old_macs':old_macs})
	result.update({'new_macs':new_macs})
	return result

def check_ips_in_file(file_name,ips):
	old_ips=[]
	new_ips=[]
	result={}
	found_flag=False
	for ip in ips:
		if len(ip)==1:
			for line in file(file_name):
				if str(ip[0]) in line:
					print bcolors.WARNING+"Red encontrada: "+bcolors.ENDC+bcolors.UNDERLINE+str(ip[0])+bcolors.ENDC
					found_flag=True
					break
			if found_flag:
				old_ips.append(ip)
			else:
				new_ips.append(ip)
		elif len(ip)>1:
			for i in ip:
				for line in file(file_name):
					if str(i) in line:
						print bcolors.WARNING+"Red encontrada: "+bcolors.ENDC+bcolors.UNDERLINE+str(i)+bcolors.ENDC
						found_flag=True
						break
				if found_flag:
					old_ips.append(i)
				else:
					new_ips.append(i)
	result.update({'old_ips':old_ips})
	result.update({'new_ips':new_ips})
	return result

def normalize(subnormal):
	normal=[]
	for element in subnormal:
		if len(element)==1 and 'list' in str(type(element)):
			normal.append(str(element[0]))
		elif len(element)>1 and 'list' in str(type(element)):
			for e in element:
				normal.append(str(e))
		else:
			normal.append(str(element))
	return normal

def get_user_variables():
	user_variables={}
	user_variables.update({'db_file':raw_input('Escribe el nombre del archivo: ')})
	user_variables.update({'db_sheet':raw_input('Escribe el nombre de la hoja en que se encuentra la informacion: ')})
	return user_variables
	#importante revisar esto

def get_line_numbers(file_name,search_array):
	lines=[]
	for element in search_array:
		line_number=1
		for line in file(file_name):
			if element in line:
				lines.append(line_number)
				break
			else:
				line_number+=1
	return lines

def get_lines(file_name):
	lines=0
	for line in file(file_name):
		lines=+1
	return lines

def get_dhcp_leases(data,networks):
	i=2
	all_devices={}
	while i<data.max_row:
		site_devices=[]
		current_inegi=data.cell(i,mapColNum('D')).value
		next_inegi=data.cell(i+1,mapColNum('D')).value
		current_device=data.cell(i,mapColNum('C')).value
		site_devices.append(current_device)
		while current_inegi==next_inegi:
			i+=1
			current_inegi=data.cell(i,mapColNum('D')).value
			next_inegi=data.cell(i+1,mapColNum('D')).value
			current_device=data.cell(i,mapColNum('C')).value
			site_devices.append(current_device)
		all_devices[current_inegi]=site_devices
		i+=1
	all_leases={}
	for key in all_devices.keys():
		site_network=list(v for k,v in networks.iteritems() if str(key) in k)
		site_leases=[]
		if len(site_network)==1:
			ip_network=ipaddress.IPv4Network(site_network[0]+'/28')
			site_devices=all_devices[key]
			i=0
			while i<len(site_devices):
				lease=ip_network[i+2],site_devices[i]
				site_leases.append(lease)
				i+=1
			all_leases.update({ip_network:site_leases})
	return all_leases

def iscDHCPformat(leases):
	all_entries=""
	for key in leases.keys():
		#print key
		subnet=ipaddress.IPv4Network(key)
		subnet_address=subnet.network_address
		netmask=subnet.netmask
		gateway=list(subnet.hosts())[0]
		broadcast=subnet.broadcast_address
		entry_data=[subnet_address,netmask,gateway,broadcast]
		hosts=[]
		for i in range(len(leases[key])):
			#print leases[key][i]
			host_entry="\thost "+leases[key][i][1]+" {hardware ethernet %(mac_address)s; fixed-address %(ip_address)s;}"%{"mac_address":nextMAC(formatMACaddress(leases[key][i][1])),"ip_address":leases[key][i][0]}
			hosts.append(host_entry)
		group_entry="{"
		for j in range(len(hosts)):
			group_entry=group_entry+"\n\t"+hosts[j-1]
			if j==len(hosts)-1:
				group_entry=group_entry+"\n\t}\n"
		entry_string="subnet %(subnet)s netmask %(netmask)s { \n\toption routers %(gateway)s;\n\toption broadcast-address %(broadcast)s;\n\toption domain-name-servers 189.247.96.1, 189.247.97.1;\n\toption subnet-mask %(netmask)s;\n\tgroup %(group_entry)s"%{"subnet":subnet_address,"netmask":netmask,"gateway":gateway,"broadcast":broadcast,"group_entry":group_entry}
		entry_string=entry_string+"}\n"
		all_entries=all_entries+entry_string
	return all_entries

def update_file(file_name,update_string):
	with open(file_name, 'r') as file:
		file_string=file.read()
		position=len(file_string)
		for c in reversed(file_string):
			if c=='}':
				beginning=file_string[:position-1]
				break
			else:
				position=position-1
		new_dhcp_string=beginning+'\n'+"#actualizacion de archivo - "+time.strftime("%Y%m%d")+'\n'+'\n'+update_string+'\n'+'}'
		try:
			new_dhcp_file=open('new_dhcp_file.conf','w+')
			new_dhcp_file.write(new_dhcp_string)
			new_dhcp_file.close()
			print 'Archivo de configuracion generado exitosamente.'
			return True
		except Exception, e:
			return False

def getNetworks(db,sheet):
	wb=openpyxl.load_workbook(db)
	data=wb[sheet]
	i=2
	networks={}
	while i<data.max_row:
		#print i
		#print data.max_row
		current_site=data.cell(i,mapColNum('A')).value
		next_site=data.cell(i+1,mapColNum('A')).value
		if current_site!=next_site:
			networks[current_site]=data.cell(i,mapColNum('B')).value
		i+=1
	return networks

networks=getNetworks('ref_net.xlsx','Sheet1')
print 'Redes cargadas con exito.'
svr='189.240.5.86'
ssh_port=60122 #puerto pera servidor de produccion
user_path='/home/cwmp'
dhcp_conf_path='/etc/dhcp/dhcpd.conf'
dhcp_bak_path='/tmp/dhcp.conf.bak.'
dhcp_upload_path='/tmp/dhcp.conf.new'
user_variables=get_user_variables()
print
output=remote_copy_to_tmp_dir(svr,ssh_port,dhcp_conf_path,user_path+dhcp_bak_path)
#output[2] es stderr entonces la siguiente condicion revisa si hay error
if len(output[2].readlines())>1:
	print bcolors.FAIL+'Error en la ejecucion del comando en el servidor remoto.'+bcolors.ENDC
	sys.exit()
#si no hay error
else:
	if ftp_download(svr,dhcp_bak_path+time.strftime("%Y%m%d"),'dhcp_file.conf.'+time.strftime("%Y%m%d")):
		data=getData(user_variables['db_file'],user_variables['db_sheet'])
		i=2
		macs=[]
		while i<data.max_row:
			macs.append(data.cell(i,mapColNum('C')).value)
			i+=1
		j=2
		inegis=[]
		while j<data.max_row:
			current_inegi=data.cell(j,mapColNum('D')).value
			next_inegi=data.cell(j+1,mapColNum('D')).value
			if current_inegi!=next_inegi:
				inegis.append(current_inegi)
			j+=1
		ip_networks=[]
		for inegi in inegis:
			site_network=list(v for k,v in networks.iteritems() if str(inegi) in k)
			ip_networks.append(site_network)
		ip_report=check_ips_in_file('dhcp_file.conf.'+time.strftime("%Y%m%d"),ip_networks)
		mac_report=check_macs_in_file('dhcp_file.conf.'+time.strftime("%Y%m%d"),macs)
		#print mac_report['old_macs']
		if len(mac_report['old_macs'])>=1:
			print '\n'+bcolors.WARNING+'ATENCION: '+bcolors.ENDC+'Es necesario comentar las siguientes lineas con MACs repetidas:'
			for line_number in get_line_numbers('dhcp_file.conf.'+time.strftime("%Y%m%d"),mac_report['old_macs']):
				print line_number
		#normalizar ip_report
		ip_report_old=normalize(ip_report['old_ips'])
		#print ip_report_old
		if len(ip_report_old)>=1:
			print '\n'+bcolors.WARNING+'ATENCION: '+bcolors.ENDC+'Es necesario deshacerse de los subnets en estas lineas:'
			for line_number in get_line_numbers('dhcp_file.conf.'+time.strftime("%Y%m%d"),ip_report_old):
				print line_number
		print 'Actualizando archivo...'
		update_file('dhcp_file.conf.'+time.strftime("%Y%m%d"),iscDHCPformat(get_dhcp_leases(data,networks)))
		if ftp_upload(svr,'new_dhcp_file.conf',user_path+dhcp_upload_path):
			print '\n'+bcolors.WARNING+'***IMPORTANTE***'+bcolors.ENDC
			print 'Ingresa al servidor remoto y ejecuta los siguientes comandos:\n'
			print 'sudo cp ~/tmp/dhcp.conf.new /etc/dhcp/dhcpd.conf'
			print 'sudo systemctl restart isc-dhcp-server'
			print '\nAl finalizar revisa el estado del servidor con systemctl status isc-dhcp-server.'
			print bcolors.WARNING+'****************'+bcolors.ENDC
		else:
			print 'No se pudo cargar el archvio en el servidor remoto.'
			print 'Favor de realizar la operacion manualmente.'

